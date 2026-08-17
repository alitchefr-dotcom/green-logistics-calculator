import io
import openpyxl
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------
# Page Configuration
# ---------------------------------------------------------
st.set_page_config(
    page_title="Green-Logistics Customs & Landed Cost Calculator",
    page_icon="⚡",
    layout="wide",
)

# ---------------------------------------------------------
# Constants & Reference Data
# ---------------------------------------------------------
SEA_EMISSION_FACTOR_G_PER_TKM = 10.0  # g CO2e / tonne-km
ROAD_EMISSION_FACTOR_G_PER_TKM = 62.0  # g CO2e / tonne-km

# ---------------------------------------------------------
# Language Selector & Translations
# ---------------------------------------------------------
lang = st.sidebar.radio("🌐 Language / שפה", ["עברית", "English"])
is_heb = lang == "עברית"

# Inject RTL style if Hebrew selected
if is_heb:
    st.markdown(
        """
        <style>
        .stApp {
            direction: rtl;
            text-align: right;
        }
        .stMarkdown, .stText, .stCaption {
            text-align: right;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

# Bilingual Dictionaries
t = {
    "title": (
        "⚡ מחשבון עלויות יבוא, מכס ו-Green Landed Cost"
        if is_heb
        else "⚡ Green-Logistics Customs & Landed Cost Calculator"
    ),
    "subtitle": (
        "מחשבון עלויות יבוא, מכס, רגולציה, אחסנה חיצונית, מחזור (EPR), פליטות פחמן (CO₂e)"
        " ו-Landed Cost לתשתיות אנרגיה (**BESS**, **MVS**, פאנלים סולאריים,"
        " ממירים ושנאים)."
        if is_heb
        else (
            "Import Duty, Tax, External Storage, Battery Recycling (EPR), Carbon Emissions"
            " (CO₂e) & Landed Cost Calculator for Renewable Energy"
            " Infrastructure (**BESS**, **MVS Skids**, Solar Panels, Inverters &"
            " Transformers)."
        )
    ),
    "route_header": (
        "🌍 מסלול השינוע והנמלים" if is_heb else "🌍 Shipping Route & Ports"
    ),
    "origin": "מדינת מוצא" if is_heb else "Origin Country",
    "origin_custom": (
        "הזן מדינת מוצא ידנית" if is_heb else "Enter Custom Origin Country"
    ),
    "port": "נמל פקידה / שחרור" if is_heb else "Port of Discharge",
    "port_custom": "הזן שם נמל ידני" if is_heb else "Enter Custom Port Name",
    "dest": "מדינת יעד סופית" if is_heb else "Final Destination Country",
    "final_dest": (
        "יעד מסירה סופי באתר (שם/קואורדינטות)"
        if is_heb
        else "Final Delivery Location (Site Name / GPS)"
    ),
    "currency_header": (
        "🔱 מטבע מקומי ושערי חליפין"
        if is_heb
        else "🔱 Local Currency & Exchange Rates"
    ),
    "ex_rate_label": (
        "שער המרה מדולר ($) למטבע מקומי"
        if is_heb
        else "Exchange Rate (USD to Local Currency)"
    ),
    "cargo_header": (
        "📋 פרטי המטען, מכולות ויחידות ציוד"
        if is_heb
        else "📋 Cargo, Containers & Equipment Units"
    ),
    "equipment": "סוג הציוד" if is_heb else "Equipment Category",
    "container_type": (
        "סוג מכולה / סיווג בטיחות"
        if is_heb
        else "Container Type & Safety Classification"
    ),
    "hs": "קוד מכס מוצע (HS Code)" if is_heb else "Suggested HS Code",
    "num_containers": "כמות מכולות" if is_heb else "Number of Containers",
    "bess_capacity": (
        "קיבולת BESS כוללת (MWh)"
        if is_heb
        else "Total BESS System Capacity (MWh)"
    ),
    "cargo_units": (
        'סה"כ יחידות ציוד (פאנלים/ממירים)'
        if is_heb
        else "Total Cargo Units (Panels / Inverters)"
    ),
    "weight": (
        "משקל ברוטו למכולה (טון)"
        if is_heb
        else "Gross Weight per Container (Tonnes)"
    ),
    "cargo_val": (
        "ערך הסחורה במקור ($ USD)"
        if is_heb
        else "Cargo FOB / EXW Value ($ USD)"
    ),
    "freight_header": (
        "🚢 עלויות הובלה וביטוח ימי"
        if is_heb
        else "🚢 Ocean Freight & Insurance"
    ),
    "freight": (
        "הובלה ימית ראשית ($ USD)" if is_heb else "Main Ocean Freight ($ USD)"
    ),
    "ins_basis": "בסיס חישוב ביטוח" if is_heb else "Insurance Calculation Basis",
    "insurance_rate": (
        "שיעור ביטוח ימי (%)" if is_heb else "Marine Insurance Rate (%)"
    ),
    "origin_exp": (
        "הוצאות במקור ($ USD)" if is_heb else "Origin Local Expenses ($ USD)"
    ),
    "tax_header": (
        "🏛️ מכס, הסכמי סחר ומיסוי"
        if is_heb
        else "🏛️ Customs, Duties & Taxation"
    ),
    "customs_rate": (
        "שיעור מכס מוערך (%)" if is_heb else "Estimated Customs Duty Rate (%)"
    ),
    "duty_incentive": "הטבת מכס / פטור" if is_heb else "Customs Duty Incentive",
    "pref_rate": (
        "שיעור מכס מועדף מאושר (%)"
        if is_heb
        else "Verified Preferential Duty Rate (%)"
    ),
    "vat": (
        'שיעור מע"מ מקומי במדינת השחרור (%)'
        if is_heb
        else "Local VAT Rate at Port of Clearance (%)"
    ),
    "port_header": (
        "⚓ עלויות נמל, THC ועמילות בארץ היעד"
        if is_heb
        else "⚓ Destination Port, THC & Brokerage"
    ),
    "port_fees": (
        "אגרות נמל וסדרנות ($ USD)"
        if is_heb
        else "Port Wharfage & Handling ($ USD)"
    ),
    "thc": (
        "דמי טיפול במסוף (THC) למכולה ($ USD)"
        if is_heb
        else "Terminal Handling Charge (THC) per Container ($ USD)"
    ),
    "brokerage": (
        "עמילות מכס ואישורים ($ USD)"
        if is_heb
        else "Customs Brokerage & Permits ($ USD)"
    ),
    "inland_per_unit": (
        "הובלה יבשתית למכולה ($ USD)"
        if is_heb
        else "Special Heavy Inland Haulage per Container ($ USD)"
    ),
    "storage_header": (
        "📦 עלויות אחסנה חיצונית ושטחי היערכות (Staging Yard)"
        if is_heb
        else "📦 External Storage & Staging Yard Costs"
    ),
    "demurrage_header": (
        "⏱️ ניהול סיכוני השהיה בנמל (Demurrage Risk)"
        if is_heb
        else "⏱️ Port Demurrage Risk Estimator"
    ),
    "free_days": (
        "ימים חופשיים בנמל (Free Days)"
        if is_heb
        else "Port Free Days Included"
    ),
    "est_port_days": (
        "ימי השהיה בפועל בנמל (משוער)"
        if is_heb
        else "Estimated Actual Port Dwell Days"
    ),
    "demurrage_rate": (
        "עלות השהיה יומית למכולה ($/Day)"
        if is_heb
        else "Daily Demurrage Rate per Container ($/Day)"
    ),
    "recycling_header": (
        "♻️ מחזור סוללות ואחריות יצרן (EPR Provision)"
        if is_heb
        else "♻️ Battery Recycling & EPR Provision"
    ),
    "recycling_mode": "מצב חישוב מחזור" if is_heb else "Recycling Calculation Mode",
    "recycling_rate": (
        "הפרשת מחזור סוף חיים ($/kWh)"
        if is_heb
        else "End-of-Life Provision ($/kWh)"
    ),
    "green_header": (
        "🌱 מדדי קיימות ופליטות פחמן (CO₂e)"
        if is_heb
        else "🌱 Sustainability & Carbon Footprint (CO₂e)"
    ),
    "sea_dist": "מרחק שינוע ימי (ק\"מ)" if is_heb else "Ocean Distance (km)",
    "land_dist": "מרחק שינוע יבשתי (ק\"מ)" if is_heb else "Inland Distance (km)",
    "cif_metric": "ערך CIF כולל" if is_heb else "Total CIF Value",
    "duty_metric": "תשלום מכס אפקטיבי" if is_heb else "Effective Duty",
    "vat_metric": 'מע"מ לתשלום' if is_heb else "VAT Payable",
    "base_landed_metric": (
        "עלות Landed Cost בסיסית" if is_heb else "Base Landed Cost"
    ),
    "risk_adj_metric": (
        'סה"כ מותאם סיכון וסוף חיים' if is_heb else "Risk-Adjusted Total"
    ),
    "breakdown_title": (
        "📊 ניתוח הצטברות עלויות (Waterfall Chart)"
        if is_heb
        else "📊 Cost Accumulation (Waterfall Chart)"
    ),
    "summary_title": (
        "📑 טבלת סיכום שלבי החישוב" if is_heb else "📑 Step-by-Step Summary Table"
    ),
    "export_title": "📥 ייצוא נתונים" if is_heb else "📥 Export Data",
    "btn_excel": (
        "📊 צור והורד דוח Excel מפורט"
        if is_heb
        else "📊 Generate Detailed Excel Report"
    ),
}

# ---------------------------------------------------------
# Sidebar Inputs
# ---------------------------------------------------------
st.sidebar.header(t["route_header"])

origin_selection = st.sidebar.selectbox(
    t["origin"],
    ["China", "Germany", "USA", "India", "Japan", "South Korea", "Other / אחר"],
    index=0,
)

if origin_selection == "Other / אחר":
    origin_country = st.sidebar.text_input(t["origin_custom"], value="Vietnam")
else:
    origin_country = origin_selection

port_defaults = {
    "Port of Haifa / Bayport (Israel)": {
        "country": "Israel",
        "dest": "Israel",
        "vat": 18.0,
        "curr": "₪ ILS",
        "sym": "₪",
        "rate": 3.70,
        "site": "Carmiel Industrial Zone / GPS: 32.9199, 35.2901",
        "sea_km": 15000,
        "land_km": 60,
    },
    "Port of Ashdod (Israel)": {
        "country": "Israel",
        "dest": "Israel",
        "vat": 18.0,
        "curr": "₪ ILS",
        "sym": "₪",
        "rate": 3.70,
        "site": "Carmiel Industrial Zone / GPS: 32.9199, 35.2901",
        "sea_km": 15200,
        "land_km": 170,
    },
    "Port of Burgas (Bulgaria)": {
        "country": "Bulgaria",
        "dest": "Bulgaria / Transit to Romania",
        "vat": 20.0,
        "curr": "€ EUR",
        "sym": "€",
        "rate": 0.92,
        "site": "Iepuresti Solar Plant, Romania",
        "sea_km": 14200,
        "land_km": 320,
    },
    "Port of Constanta (Romania)": {
        "country": "Romania",
        "dest": "Romania",
        "vat": 19.0,
        "curr": "€ EUR",
        "sym": "€",
        "rate": 0.92,
        "site": "Ghimpati Solar Plant, Romania",
        "sea_km": 14500,
        "land_km": 280,
    },
    "Port of Piraeus (Greece)": {
        "country": "Greece",
        "dest": "Greece",
        "vat": 24.0,
        "curr": "€ EUR",
        "sym": "€",
        "rate": 0.92,
        "site": "Athens Substation Site",
        "sea_km": 13800,
        "land_km": 40,
    },
    "Port of Thessaloniki (Greece)": {
        "country": "Greece",
        "dest": "Greece",
        "vat": 24.0,
        "curr": "€ EUR",
        "sym": "€",
        "rate": 0.92,
        "site": "Northern Greece Substation",
        "sea_km": 14000,
        "land_km": 80,
    },
    "Port of Rauma (Finland)": {
        "country": "Finland",
        "dest": "Finland",
        "vat": 25.5,
        "curr": "€ EUR",
        "sym": "€",
        "rate": 0.92,
        "site": "Tuovila BESS Site, Finland",
        "sea_km": 18500,
        "land_km": 210,
    },
    "Port of Vuosaari / Helsinki (Finland)": {
        "country": "Finland",
        "dest": "Finland",
        "vat": 25.5,
        "curr": "€ EUR",
        "sym": "€",
        "rate": 0.92,
        "site": "Pyhäsalmi BESS Site, Finland",
        "sea_km": 18300,
        "land_km": 450,
    },
    "Port of Rotterdam (Netherlands)": {
        "country": "Netherlands",
        "dest": "Netherlands / EU",
        "vat": 21.0,
        "curr": "€ EUR",
        "sym": "€",
        "rate": 0.92,
        "site": "EU Central Hub",
        "sea_km": 18000,
        "land_km": 100,
    },
    "Port of Antwerp-Bruges (Belgium)": {
        "country": "Belgium",
        "dest": "Belgium / EU",
        "vat": 21.0,
        "curr": "€ EUR",
        "sym": "€",
        "rate": 0.92,
        "site": "EU Distribution Center",
        "sea_km": 18100,
        "land_km": 120,
    },
    "Other Port": {
        "country": "Other",
        "dest": "",
        "vat": 0.0,
        "curr": "$ USD",
        "sym": "$",
        "rate": 1.00,
        "site": "",
        "sea_km": 0,
        "land_km": 0,
    },
}

selected_port_key = st.sidebar.selectbox(
    t["port"], list(port_defaults.keys()), index=0
)

if selected_port_key == "Other Port":
    selected_port = st.sidebar.text_input(
        t["port_custom"], value="Port of Limassol"
    )
    custom_dest_country = st.sidebar.text_input(t["dest"], value="Cyprus")
    custom_vat = st.sidebar.number_input(t["vat"], min_value=0.0, value=19.0)
    custom_curr = st.sidebar.selectbox("Currency", ["€ EUR", "$ USD", "Other"])
    custom_sym = "€" if "EUR" in custom_curr else "$"
    custom_ex_rate = st.sidebar.number_input(
        "USD Exchange Rate", min_value=0.01, value=0.92
    )
    dest_info = {
        "country": "Custom",
        "dest": custom_dest_country,
        "vat": custom_vat,
        "curr": custom_curr,
        "sym": custom_sym,
        "rate": custom_ex_rate,
        "site": "",
        "sea_km": 14000,
        "land_km": 100,
    }
else:
    selected_port = selected_port_key
    dest_info = port_defaults[selected_port_key]

dest_country = st.sidebar.text_input(t["dest"], value=dest_info["dest"])

final_destination = st.sidebar.text_input(
    t["final_dest"],
    value=dest_info["site"],
    key=f"final_site_{selected_port_key}",
)

port_base_country = dest_info["country"]
if (
    port_base_country != "Custom"
    and port_base_country.lower() not in dest_country.lower()
):
    st.sidebar.warning(
        f"⚠️ **Route Alert:** Selected port ({selected_port}) country differs"
        f" from final destination ({dest_country}). This will be treated as a"
        " transit / cross-border shipment scenario."
    )

st.sidebar.divider()
st.sidebar.header(t["currency_header"])

curr_symbol = dest_info["sym"]
ex_rate = st.sidebar.number_input(
    f"{t['ex_rate_label']} (1 USD ➔ {dest_info['curr']})",
    min_value=0.01,
    value=float(dest_info["rate"]),
    step=0.01,
)

st.sidebar.divider()
st.sidebar.header(t["cargo_header"])

equipment_type = st.sidebar.selectbox(
    t["equipment"],
    [
        "BESS Container (Battery Energy Storage Systems - DG Class 9)",
        "MVS - Medium Voltage Stations / Skids (Non-Hazmat / Non-DG)",
        "MVS Accessories & Switchgear (Non-Hazmat)",
        "Solar Panels (PV Modules)",
        "Inverters & Transformers",
        "Wind Turbine Components",
    ],
)

is_bess = "BESS" in equipment_type
is_mvs = "MVS" in equipment_type
is_solar = "Solar" in equipment_type

if is_bess:
    container_options = {
        "20HC SOC BESS (DG Class 9 / Hazmat)": 520.0,
        "40ft / 40HC SOC BESS (DG Class 9 / Hazmat)": 580.0,
    }
    }
elif is_mvs:
    container_options = {
        "20ft SOC MVS Skid (Non-Hazmat / Standard Cargo)": 350.0,
        "40ft / 40HC SOC MVS Skid (Non-Hazmat / Standard Cargo)": 450.0,
        "Special Equipment MVS (Flat Rack / Open Top)": 480.0,
    }
else:
    container_options = {
        "40ft / 40HC Standard Dry Container": 280.0,
        "Special Equipment (Open Top / Flat Rack)": 480.0,
    }

container_type = st.sidebar.selectbox(
    t["container_type"], list(container_options.keys()), index=0
)
default_thc_value = container_options[container_type]

hs_defaults = {
    "BESS Container (Battery Energy Storage Systems - DG Class 9)": "8507.60",
    "MVS - Medium Voltage Stations / Skids (Non-Hazmat / Non-DG)": "8504.22",
    "MVS Accessories & Switchgear (Non-Hazmat)": "8537.20",
    "Solar Panels (PV Modules)": "8541.43",
    "Inverters & Transformers": "8504.40",
    "Wind Turbine Components": "8502.31",
}

eu_mfn_customs_rates = {
    "BESS Container (Battery Energy Storage Systems - DG Class 9)": 2.7,
    "MVS - Medium Voltage Stations / Skids (Non-Hazmat / Non-DG)": 2.1,
    "MVS Accessories & Switchgear (Non-Hazmat)": 2.1,
    "Solar Panels (PV Modules)": 0.0,
    "Inverters & Transformers": 2.1,
    "Wind Turbine Components": 2.7,
}

hs_code = st.sidebar.text_input(
    t["hs"], value=hs_defaults.get(equipment_type, "8507.60")
)
st.sidebar.caption("💡 *Suggested HS Code — user verification required.*")

num_containers = st.sidebar.number_input(
    t["num_containers"], min_value=1, value=5, step=1
)

if is_bess:
    bess_mwh_capacity = st.sidebar.number_input(
        t["bess_capacity"], min_value=0.1, value=10.0, step=0.5
    )
    total_cargo_units = int(bess_mwh_capacity * 1000)  # kWh
else:
    bess_mwh_capacity = 0.0
    default_units_per_container = 1 if is_mvs else (600 if is_solar else 20)
    total_cargo_units = st.sidebar.number_input(
        t["cargo_units"],
        min_value=1,
        value=int(num_containers * default_units_per_container),
        step=10,
    )

container_weight = st.sidebar.number_input(
    t["weight"],
    min_value=1.0,
    max_value=100.0,
    value=30.0 if is_mvs else 45.0,
    step=1.0,
)
total_weight = container_weight * num_containers
cargo_value_usd = st.sidebar.number_input(
    t["cargo_val"], min_value=0.0, value=150000.0, step=1000.0
)

st.sidebar.divider()
st.sidebar.header(t["freight_header"])
freight_cost_usd = st.sidebar.number_input(
    t["freight"], min_value=0.0, value=12000.0, step=500.0
)
origin_expenses_usd = st.sidebar.number_input(
    t["origin_exp"], min_value=0.0, value=1500.0, step=100.0
)

insurance_basis = st.sidebar.selectbox(
    t["ins_basis"],
    ["FOB Cargo Value Only", "CIF Base (FOB + Freight + 10% Uplift)"],
    index=1,
)

insurance_rate = (
    st.sidebar.number_input(
        t["insurance_rate"], min_value=0.0, value=0.3, step=0.05
    )
    / 100
)

st.sidebar.divider()
st.sidebar.header(t["tax_header"])

default_customs_suggested = (
    0.0
    if (
        "Israel" in dest_country
        or "Haifa" in selected_port
        or "Ashdod" in selected_port
    )
    else eu_mfn_customs_rates.get(equipment_type, 2.7)
)

customs_rate_input = st.sidebar.number_input(
    t["customs_rate"],
    min_value=0.0,
    max_value=100.0,
    value=float(default_customs_suggested),
    step=0.1,
)

duty_incentive = st.sidebar.selectbox(
    t["duty_incentive"],
    [
        "Standard MFN Duty Rate",
        "Preferential FTA Treatment (Verification Required)",
        "Green Incentive (50% Duty Reduction)",
        "Full Regulatory Exemption (0% Duty)",
    ],
    index=0,
)

if "Preferential FTA" in duty_incentive:
    verified_pref_rate = st.sidebar.number_input(
        t["pref_rate"], min_value=0.0, max_value=100.0, value=0.0, step=0.1
    )
else:
    verified_pref_rate = 0.0

vat_rate = (
    st.sidebar.number_input(
        t["vat"], min_value=0.0, value=float(dest_info["vat"]), step=0.5
    )
    / 100
)

st.sidebar.divider()
st.sidebar.header(t["port_header"])
port_fees_usd = st.sidebar.number_input(
    t["port_fees"], min_value=0.0, value=1200.0, step=50.0
)

thc_fees_usd = st.sidebar.number_input(
    t["thc"],
    min_value=0.0,
    value=float(default_thc_value),
    step=25.0,
    key=f"thc_{container_type}",
)
total_thc_usd = thc_fees_usd * num_containers

brokerage_fees_usd = st.sidebar.number_input(
    t["brokerage"], min_value=0.0, value=850.0, step=50.0
)

inland_per_container_usd = st.sidebar.number_input(
    t["inland_per_unit"], min_value=0.0, value=900.0, step=50.0
)
total_inland_transport_usd = inland_per_container_usd * num_containers

# ---------------------------------------------------------
# External Storage Component (Staging Yard)
# ---------------------------------------------------------
st.sidebar.divider()
st.sidebar.header(t["storage_header"])

enable_storage = st.sidebar.checkbox(
    "כלול אחסנה חיצונית?" if is_heb else "Include External Storage?", value=False
)

if enable_storage:
    storage_site = st.sidebar.selectbox(
        "מיקום החצר" if is_heb else "Storage Location",
        [
            "חצר פתוחה קרובה לנמל (Near Port Yard)",
            "חצר היערכות קרובה לאתר (Near Site Staging)",
            "אחסנה ייעודית לחומ\"ס (DG Yard)",
        ],
    )
    storage_days = st.sidebar.number_input(
        "מספר ימי אחסנה" if is_heb else "Storage Duration (Days)",
        min_value=1,
        value=14,
        step=1,
    )
    storage_daily_rate = st.sidebar.number_input(
        "תעריף יומי למכולה ($)" if is_heb else "Daily Rate per Container ($)",
        min_value=0.0,
        value=18.0,
        step=2.0,
    )
    storage_handling_fee = st.sidebar.number_input(
        "דמי טיפול כניסה/יציאה ($/מכולה)" if is_heb else "Handling Fee Gate In/Out ($)",
        min_value=0.0,
        value=120.0,
        step=10.0,
    )
    storage_drayage_fee = st.sidebar.number_input(
        "הובלה מקשרת לחצר ($/מכולה)" if is_heb else "Short Drayage to Yard ($)",
        min_value=0.0,
        value=250.0,
        step=25.0,
    )

    total_storage_daily = storage_days * storage_daily_rate * num_containers
    total_storage_handling = storage_handling_fee * num_containers
    total_storage_drayage = storage_drayage_fee * num_containers

    total_external_storage_usd = (
        total_storage_daily + total_storage_handling + total_storage_drayage
    )
else:
    storage_site = "N/A"
    storage_days = 0
    total_external_storage_usd = 0.0

st.sidebar.divider()
st.sidebar.header(t["demurrage_header"])
free_days = st.sidebar.number_input(t["free_days"], min_value=0, value=14, step=1)
est_port_days = st.sidebar.number_input(
    t["est_port_days"], min_value=0, value=16, step=1
)
demurrage_daily_rate = st.sidebar.number_input(
    t["demurrage_rate"], min_value=0.0, value=120.0, step=10.0
)

st.sidebar.divider()
st.sidebar.header(t["recycling_header"])

recycling_mode = st.sidebar.selectbox(
    t["recycling_mode"],
    [
        "Not Included",
        "Internal End-of-Life Financial Provision",
        "Verified Regulatory EPR Obligation (EU 2023/1542)",
    ],
    index=1 if is_bess else 0,
)

if is_bess and recycling_mode != "Not Included":
    recycling_rate_per_kwh = st.sidebar.number_input(
        t["recycling_rate"], min_value=0.0, value=12.0, step=1.0
    )
    total_recycling_provision_usd = (
        bess_mwh_capacity * 1000 * recycling_rate_per_kwh
    )
else:
    recycling_rate_per_kwh = 0.0
    total_recycling_provision_usd = 0.0

st.sidebar.divider()
st.sidebar.header(t["green_header"])
sea_dist_km = st.sidebar.number_input(
    t["sea_dist"], min_value=0, value=int(dest_info["sea_km"]), step=500
)
land_dist_km = st.sidebar.number_input(
    t["land_dist"], min_value=0, value=int(dest_info["land_km"]), step=10
)

# ---------------------------------------------------------
# Pure Core Math Function
# ---------------------------------------------------------
def calculate_landed_cost():
    if insurance_basis == "FOB Cargo Value Only":
        insurance = cargo_value_usd * insurance_rate
    else:
        insurance = (
            (cargo_value_usd + freight_cost_usd + origin_expenses_usd)
            * 1.10
            * insurance_rate
        )

    cif = cargo_value_usd + freight_cost_usd + origin_expenses_usd + insurance

    if "Full Regulatory Exemption" in duty_incentive:
        eff_duty_rate = 0.0
    elif "Preferential FTA" in duty_incentive:
        eff_duty_rate = verified_pref_rate / 100
    elif "Green Incentive (50%" in duty_incentive:
        eff_duty_rate = (customs_rate_input / 100) * 0.5
    else:
        eff_duty_rate = customs_rate_input / 100

    duty = cif * eff_duty_rate
    vat_base = cif + duty + port_fees_usd
    vat = vat_base * vat_rate

    local_clearance = (
        port_fees_usd
        + total_thc_usd
        + brokerage_fees_usd
        + total_inland_transport_usd
        + total_external_storage_usd
    )
    base_landed_net = cif + duty + local_clearance

    demurrage_days = max(0, est_port_days - free_days)
    demurrage = demurrage_days * demurrage_daily_rate * num_containers
    risk_adjusted_total = (
        base_landed_net + demurrage + total_recycling_provision_usd
    )

    # CO2e Calculation
    sea_co2 = (
        total_weight * sea_dist_km * SEA_EMISSION_FACTOR_G_PER_TKM
    ) / 1_000_000
    road_co2 = (
        total_weight * land_dist_km * ROAD_EMISSION_FACTOR_G_PER_TKM
    ) / 1_000_000
    total_co2 = sea_co2 + road_co2

    return {
        "insurance": insurance,
        "cif": cif,
        "duty_rate": eff_duty_rate,
        "duty": duty,
        "vat": vat,
        "local_clearance": local_clearance,
        "base_landed_net": base_landed_net,
        "demurrage_days": demurrage_days,
        "demurrage": demurrage,
        "risk_adjusted_total": risk_adjusted_total,
        "sea_co2": sea_co2,
        "road_co2": road_co2,
        "total_co2": total_co2,
    }


res = calculate_landed_cost()

# Local Currency Conversion
cif_loc = res["cif"] * ex_rate
duty_loc = res["duty"] * ex_rate
vat_loc = res["vat"] * ex_rate
base_landed_net_loc = res["base_landed_net"] * ex_rate
risk_adjusted_loc = res["risk_adjusted_total"] * ex_rate

# KPI Calculation
if is_bess and bess_mwh_capacity > 0:
    kpi_label = "Base Landed Cost / MWh"
    kpi_value_usd = res["base_landed_net"] / bess_mwh_capacity
elif is_solar and total_cargo_units > 0:
    kpi_label = "Base Landed Cost / Panel"
    kpi_value_usd = res["base_landed_net"] / total_cargo_units
else:
    kpi_label = "Base Landed Cost / Container"
    kpi_value_usd = res["base_landed_net"] / num_containers

kpi_value_loc = kpi_value_usd * ex_rate

# ---------------------------------------------------------
# UI Display & Metric Cards
# ---------------------------------------------------------
site_display = final_destination if final_destination else "N/A"

st.title(t["title"])
st.caption(t["subtitle"])

if is_heb:
    st.info(
        f"📍 **מסלול:** מ-**{origin_country}** דרך **{selected_port}** ➔"
        f" **{dest_country}** | **ציוד:** `{equipment_type}` | **כמות:**"
        f" `{num_containers}` מכולות | **פליטת פחמן משוערת:**"
        f" `{res['total_co2']:.2f} Ton CO₂e`"
    )
else:
    st.info(
        f"📍 **Route:** From **{origin_country}** via **{selected_port}** ➔"
        f" **{dest_country}** | **Cargo:** `{equipment_type}` | **Volume:**"
        f" `{num_containers}` Containers | **Est. Transport CO₂e:**"
        f" `{res['total_co2']:.2f} Ton CO₂e`"
    )

# China Export Regulatory Alert
if origin_country == "China" and is_bess:
    if is_heb:
        st.warning(
            "🇨🇳 **התראת רגולציית יצוא מסין (China Export DG Compliance):**\n"
            "* **אישור אריזת חומ\"ס (危包证 - Dangerous Package Certificate):** חובה לוודא מול המפעל/הספק בסין הנפקת תעודת אריזה בתוקף מה-CIQ/המכס הסיני.\n"
            "* **חלון זמן היערכות:** הנפקת האישור דורשת 10-14 ימי עסקים לפני הגעת המכולות לנמל (שנגחאי/נינגבו/צינגדאו).\n"
            "* **מסמכי חובה:** דוח UN 38.3 מעודכן ו-MSDS באנגלית וסינית."
        )
    else:
        st.warning(
            "🇨🇳 **China Export DG Regulatory Alert:**\n"
            "* **Dangerous Package Certificate (危包证):** Mandatory verification with Chinese supplier for valid CIQ export packaging certificate.\n"
            "* **Lead Time Warning:** Certificate issuance requires 10-14 business days prior to port gate-in (Shanghai/Ningbo/Qingdao).\n"
            "* **Required Documents:** Valid UN 38.3 test report and bilingual MSDS."
        )

if is_bess:
    if is_heb:
        st.warning(
            "🔥 **התראת חומרים מסוכנים ורגולציית סוללות (BESS - DG Class 9):**\n*"
            " **DG Handling:** סוללות ליתיום-יוון מוגדרות כחומר מסוכן (UN 3536)."
            " דורש DGD ואישור איחסון מסוכנים בנמל.\n* **EU Battery Passport & EPR:**"
            " באיחוד האירופי נדרש דרכון סוללה דיגיטלי והסדר מחזור סוף חיים מורשה."
        )
    else:
        st.warning(
            "🔥 **Dangerous Goods & EU Battery Passport Alert (BESS - DG Class 9):**\n*"
            " **DG Handling:** Lithium-ion batteries (UN 3536) require DGD"
            " documentation and port permits.\n* **EU Battery Passport & EPR:** EU"
            " regulations mandate a digital battery passport and EoL recycling"
            " provision."
        )

if container_weight >= 40.0:
    if is_heb:
        st.error(
            f"🚨 **התראת משקל כבד ({container_weight} טון למכולה | סה\"כ"
            f" {total_weight} טון):** מחייב היתרי הובלה חורגת (Special Transport"
            " Permits) ובדיקות עומס סרנים."
        )
    else:
        st.error(
            f"🚨 **Heavy Cargo Alert ({container_weight} Tonnes/Unit | Total:"
            f" {total_weight} Tonnes):** Requires Special Transport Permits and axle"
            " load checks."
        )

clean_hs = hs_code.replace(".", "").strip()
if (
    "Israel" in dest_country
    or "Haifa" in selected_port
    or "Ashdod" in selected_port
):
    customs_url = f"https://www.gov.il/he/departments/dynamiccollectors/customs-tariff?tariffNumber={clean_hs}"
    link_text = (
        "🔗 חיפוש בתעריף המכס הישראלי"
        if is_heb
        else "🔗 Search Israel Customs Tariff Database"
    )
else:
    customs_url = f"https://trade.ec.europa.eu/access-to-markets/en/home?product_code={clean_hs}"
    link_text = (
        f"🔗 חיפוש בפורטל Access2Markets עבור {dest_country}"
        if is_heb
        else f"🔗 Access EU Access2Markets Tariff Portal for {dest_country}"
    )

st.markdown(f"[{link_text}]({customs_url})")

col1, col2, col3, col4, col5 = st.columns(5)
col1.metric(t["cif_metric"], f"${res['cif']:,.2f}", f"{curr_symbol}{cif_loc:,.2f}")
col2.metric(
    t["duty_metric"],
    f"${res['duty']:,.2f}",
    f"{curr_symbol}{duty_loc:,.2f} ({res['duty_rate']*100:.1f}%)",
)
col3.metric(
    f"{t['vat_metric']} ({vat_rate*100:.1f}%)",
    f"${res['vat']:,.2f}",
    f"{curr_symbol}{vat_loc:,.2f}",
)
col4.metric(
    t["base_landed_metric"],
    f"${res['base_landed_net']:,.2f}",
    f"{curr_symbol}{base_landed_net_loc:,.2f}",
)
col5.metric(
    t["risk_adj_metric"],
    f"${res['risk_adjusted_total']:,.2f}",
    f"{curr_symbol}{risk_adjusted_loc:,.2f}",
)

st.divider()

left_col, right_col = st.columns([1, 1])

# ---------------------------------------------------------
# Waterfall Chart
# ---------------------------------------------------------
with left_col:
    st.subheader(t["breakdown_title"])

    x_labels = (
        [
            "ציוד FOB",
            "הובלה ימית",
            "ביטוח",
            "מכס",
            "נמל ו-THC",
            "עמילות",
            "הובלה יבשתית",
            "אחסנה חיצונית",
            "Landed Cost בסיסי",
            "סיכון השהיות",
            "הפרשת מחזור",
            'סה"כ מותאם',
        ]
        if is_heb
        else [
            "FOB Cargo",
            "Freight",
            "Insurance",
            "Customs Duty",
            "Port & THC",
            "Brokerage",
            "Inland Haulage",
            "Ext. Storage",
            "Base Landed Cost",
            "Demurrage Risk",
            "Recycling Provision",
            "Risk-Adjusted Total",
        ]
    )

    fig = go.Figure(
        go.Waterfall(
            name="Landed Cost Breakdown",
            orientation="v",
            measure=[
                "relative",
                "relative",
                "relative",
                "relative",
                "relative",
                "relative",
                "relative",
                "relative",
                "total",
                "relative",
                "relative",
                "total",
            ],
            x=x_labels,
            textposition="outside",
            text=[
                f"${cargo_value_usd:,.0f}",
                f"${freight_cost_usd+origin_expenses_usd:,.0f}",
                f"${res['insurance']:,.0f}",
                f"${res['duty']:,.0f}",
                f"${port_fees_usd+total_thc_usd:,.0f}",
                f"${brokerage_fees_usd:,.0f}",
                f"${total_inland_transport_usd:,.0f}",
                f"${total_external_storage_usd:,.0f}",
                f"${res['base_landed_net']:,.0f}",
                f"${res['demurrage']:,.0f}",
                f"${total_recycling_provision_usd:,.0f}",
                f"${res['risk_adjusted_total']:,.0f}",
            ],
            y=[
                cargo_value_usd,
                freight_cost_usd + origin_expenses_usd,
                res["insurance"],
                res["duty"],
                port_fees_usd + total_thc_usd,
                brokerage_fees_usd,
                total_inland_transport_usd,
                total_external_storage_usd,
                0,
                res["demurrage"],
                total_recycling_provision_usd,
                0,
            ],
            connector={"line": {"color": "rgb(63, 63, 63)"}},
        )
    )
    fig.update_layout(
        showlegend=False,
        height=500,
        margin=dict(l=20, r=20, t=30, b=20),
    )
    st.plotly_chart(fig, use_container_width=True)

with right_col:
    st.subheader(t["summary_title"])
    summary_df = pd.DataFrame([
        {
            "Detail": "Route / מסלול",
            "Value ($ USD)": f"{origin_country} ➔ {selected_port}",
            f"Local ({curr_symbol})": dest_info["curr"],
        },
        {
            "Detail": "Equipment Category",
            "Value ($ USD)": equipment_type,
            f"Local ({curr_symbol})": container_type,
        },
        {
            "Detail": "Volume / Capacity",
            "Value ($ USD)": f"{num_containers} Containers",
            f"Local ({curr_symbol})": (
                f"{bess_mwh_capacity} MWh" if is_bess else f"{total_cargo_units} Units"
            ),
        },
        {
            "Detail": "FOB Cargo Value",
            "Value ($ USD)": f"${cargo_value_usd:,.2f}",
            f"Local ({curr_symbol})": f"{curr_symbol}{cargo_value_usd*ex_rate:,.2f}",
        },
        {
            "Detail": "Ocean Freight + Origin",
            "Value ($ USD)": f"${freight_cost_usd + origin_expenses_usd:,.2f}",
            f"Local ({curr_symbol})": (
                f"{curr_symbol}{(freight_cost_usd + origin_expenses_usd)*ex_rate:,.2f}"
            ),
        },
        {
            "Detail": "Marine Insurance",
            "Value ($ USD)": f"${res['insurance']:,.2f}",
            f"Local ({curr_symbol})": (
                f"{curr_symbol}{res['insurance']*ex_rate:,.2f}"
            ),
        },
        {
            "Detail": "Total CIF Value",
            "Value ($ USD)": f"${res['cif']:,.2f}",
            f"Local ({curr_symbol})": f"{curr_symbol}{cif_loc:,.2f}",
        },
        {
            "Detail": "Customs Duty",
            "Value ($ USD)": f"${res['duty']:,.2f}",
            f"Local ({curr_symbol})": f"{curr_symbol}{duty_loc:,.2f}",
        },
        {
            "Detail": f"Local VAT ({vat_rate*100:.1f}%)",
            "Value ($ USD)": f"${res['vat']:,.2f}",
            f"Local ({curr_symbol})": f"{curr_symbol}{vat_loc:,.2f}",
        },
        {
            "Detail": f"Port Fees & THC ({num_containers} cont.)",
            "Value ($ USD)": f"${port_fees_usd + total_thc_usd:,.2f}",
            f"Local ({curr_symbol})": (
                f"{curr_symbol}{(port_fees_usd + total_thc_usd)*ex_rate:,.2f}"
            ),
        },
        {
            "Detail": f"Inland Heavy Haulage ({num_containers} cont.)",
            "Value ($ USD)": f"${total_inland_transport_usd:,.2f}",
            f"Local ({curr_symbol})": (
                f"{curr_symbol}{total_inland_transport_usd*ex_rate:,.2f}"
            ),
        },
        {
            "Detail": f"External Storage / Staging Yard",
            "Value ($ USD)": f"${total_external_storage_usd:,.2f}",
            f"Local ({curr_symbol})": (
                f"{curr_symbol}{total_external_storage_usd*ex_rate:,.2f}"
            ),
        },
        {
            "Detail": "🟢 Base Landed Cost (Invoice Net)",
            "Value ($ USD)": f"${res['base_landed_net']:,.2f}",
            f"Local ({curr_symbol})": f"{curr_symbol}{base_landed_net_loc:,.2f}",
        },
        {
            "Detail": f"🟡 Est. Demurrage Risk ({res['demurrage_days']} days)",
            "Value ($ USD)": f"${res['demurrage']:,.2f}",
            f"Local ({curr_symbol})": (
                f"{curr_symbol}{res['demurrage']*ex_rate:,.2f}"
            ),
        },
        {
            "Detail": f"🟡 Recycling / EPR Provision ({recycling_mode})",
            "Value ($ USD)": f"${total_recycling_provision_usd:,.2f}",
            f"Local ({curr_symbol})": (
                f"{curr_symbol}{total_recycling_provision_usd*ex_rate:,.2f}"
            ),
        },
        {
            "Detail": "🔵 Risk-Adjusted Total",
            "Value ($ USD)": f"${res['risk_adjusted_total']:,.2f}",
            f"Local ({curr_symbol})": f"{curr_symbol}{risk_adjusted_loc:,.2f}",
        },
        {
            "Detail": f"⚡ Key KPI ({kpi_label})",
            "Value ($ USD)": f"${kpi_value_usd:,.2f}",
            f"Local ({curr_symbol})": f"{curr_symbol}{kpi_value_loc:,.2f}",
        },
        {
            "Detail": "🌱 Est. Transport CO₂e Footprint",
            "Value ($ USD)": f"{res['total_co2']:.2f} Ton CO₂e",
            f"Local ({curr_symbol})": (
                f"Sea: {res['sea_co2']:.1f}t | Road: {res['road_co2']:.1f}t"
            ),
        },
    ])
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

st.divider()

# ---------------------------------------------------------
# Polished Multi-Tab Excel Generator with Formatting
# ---------------------------------------------------------
def generate_excel_bytes():
    wb = openpyxl.Workbook()

    title_font = Font(name="Calibri", size=14, bold=True, color="16A085")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="16A085", end_color="16A085", fill_type="solid"
    )
    thin_border_side = Side(border_style="thin", color="CCCCCC")
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )

    ws1 = wb.active
    ws1.title = "Landed Cost Summary"
    ws1.views.sheetView[0].showGridLines = True
    ws1.freeze_panes = "A4"

    ws1["A1"] = (
        "דוח מחשבון עלויות יבוא, פחמן ומחזור - Green Logistics"
        if is_heb
        else "Green-Logistics Customs, Carbon & Risk-Adjusted Cost Report"
    )
    ws1["A1"].font = title_font

    headers1 = [
        "רכיב עלות / פרט" if is_heb else "Cost Element / Detail",
        "סכום ($ USD)" if is_heb else "Amount ($ USD)",
        f"סכום במטבע מקומי ({curr_symbol})"
        if is_heb
        else f"Amount ({curr_symbol} Local)",
        "סיווג ורגולציה" if is_heb else "Category & Notes",
    ]
    ws1.append([])
    ws1.append(headers1)

    for col_num in range(1, 5):
        cell = ws1.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    data1 = [
        [
            "מדינת מוצא" if is_heb else "Origin Country",
            origin_country,
            origin_country,
            "Origin Location",
        ],
        [
            "סוג הציוד" if is_heb else "Equipment Category",
            equipment_type,
            equipment_type,
            "Equipment Type",
        ],
        [
            "סוג מכולה" if is_heb else "Container Type",
            container_type,
            container_type,
            "Container Category",
        ],
        [
            "כמות מכולות vs נפח" if is_heb else "Containers vs Volume",
            f"{num_containers} Containers",
            (
                f"{bess_mwh_capacity} MWh Capacity"
                if is_bess
                else f"{total_cargo_units} Equipment Units"
            ),
            "Cargo Volume",
        ],
        [
            "נמל פקידה" if is_heb else "Port of Discharge",
            selected_port,
            selected_port,
            "Discharge Port",
        ],
        [
            "מדינת יעד" if is_heb else "Final Destination Country",
            dest_country,
            dest_country,
            "Destination Country",
        ],
        [
            "ערך FOB" if is_heb else "FOB Cargo Value",
            cargo_value_usd,
            cargo_value_usd * ex_rate,
            "Invoice Base",
        ],
        [
            "הובלה ימית" if is_heb else "Ocean Freight + Origin",
            freight_cost_usd + origin_expenses_usd,
            (freight_cost_usd + origin_expenses_usd) * ex_rate,
            "Main Freight",
        ],
        [
            "ביטוח ימי" if is_heb else "Marine Insurance",
            res["insurance"],
            res["insurance"] * ex_rate,
            f"Basis: {insurance_basis}",
        ],
        [
            "ערך CIF" if is_heb else "Total CIF Value",
            res["cif"],
            cif_loc,
            "Customs Base",
        ],
        [
            "מכס אפקטיבי" if is_heb else "Customs Duty",
            res["duty"],
            duty_loc,
            f"{duty_incentive} ({res['duty_rate']*100:.1f}%)",
        ],
        [
            'מע"מ מקומי' if is_heb else "Local VAT Amount",
            res["vat"],
            vat_loc,
            f"{vat_rate*100:.1f}% Local VAT",
        ],
        [
            "אגרות נמל, THC ועמילות" if is_heb else "Port Fees, THC & Brokerage",
            port_fees_usd + total_thc_usd + brokerage_fees_usd,
            (port_fees_usd + total_thc_usd + brokerage_fees_usd) * ex_rate,
            "Port Clearance",
        ],
        [
            "הובלה יבשתית" if is_heb else "Inland Heavy Haulage",
            total_inland_transport_usd,
            total_inland_transport_usd * ex_rate,
            f"Heavy Haulage ({num_containers} cont.)",
        ],
        [
            "אחסנה חיצונית" if is_heb else "External Storage Costs",
            total_external_storage_usd,
            total_external_storage_usd * ex_rate,
            f"Yard: {storage_site} ({storage_days} days)",
        ],
        [
            "🟢 Base Landed Cost (Net)",
            res["base_landed_net"],
            base_landed_net_loc,
            "Actual Invoice Total",
        ],
        [
            "🟡 Est. Demurrage Risk",
            res["demurrage"],
            res["demurrage"] * ex_rate,
            f"{res['demurrage_days']} Excess Days Risk",
        ],
        [
            "🟡 Battery Recycling EPR Provision",
            total_recycling_provision_usd,
            total_recycling_provision_usd * ex_rate,
            f"{recycling_mode} (${recycling_rate_per_kwh}/kWh)",
        ],
        [
            "🔵 Risk-Adjusted Total",
            res["risk_adjusted_total"],
            risk_adjusted_loc,
            "Total Budget Exposure",
        ],
        [
            f"⚡ Key KPI ({kpi_label})",
            kpi_value_usd,
            kpi_value_loc,
            "Efficiency Metric",
        ],
    ]

    for row_idx, row in enumerate(data1, start=4):
        ws1.append(row)
        for col_idx in [2, 3]:
            cell = ws1.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
            cell.border = thin_border

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws2 = wb.create_sheet(title="Assumptions & Status")
    ws2.views.sheetView[0].showGridLines = True
    ws2.freeze_panes = "A4"

    ws2["A1"] = (
        "הנחות יסוד וסטטוס אימות - Green Logistics"
        if is_heb
        else "Calculation Assumptions & Verification Audit Trail"
    )
    ws2["A1"].font = title_font

    headers2 = [
        "פרמטר / רכיב" if is_heb else "Parameter / Component",
        "ערך / הגדרה" if is_heb else "Configured Value",
        "סטטוס אימות" if is_heb else "Verification Status",
        "מקור / הערות" if is_heb else "Source & Notes",
    ]
    ws2.append([])
    ws2.append(headers2)

    for col_num in range(1, 5):
        cell = ws2.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    data2 = [
        ["HS Code Classification", hs_code, "🟡 Estimated / Suggested", "Suggested classification — requires broker check"],
        ["Official Duty Rate", f"{customs_rate_input}%", "🟡 Estimated", "Base MFN Tariff rate"],
        ["Duty Incentive Status", duty_incentive, "🔴 User Confirmed", "Requires valid Certificate of Origin (COO)"],
        ["Local VAT Rate", f"{vat_rate*100:.1f}%", "🟢 Verified", f"Port of clearance standard VAT rate for {dest_country}"],
        ["China Export Compliance", "Dangerous Goods (危包证)", "⚠️ Required for China", "CIQ Dangerous Package Cert & UN 38.3 test"],
        ["External Storage Yard", storage_site, "🟡 Configured", f"{storage_days} Days storage included in estimate"],
        ["Sea Distance Factor", f"{sea_dist_km} km", "🟡 Estimated", "Estimated maritime route distance"],
        ["Sea CO2e Factor", f"{SEA_EMISSION_FACTOR_G_PER_TKM} g/t-km", "🟢 Configured", "GLEC Framework maritime factor estimate"],
        ["Road CO2e Factor", f"{ROAD_EMISSION_FACTOR_G_PER_TKM} g/t-km", "🟢 Configured", "GLEC Framework road transport estimate"],
        ["Demurrage Free Days", f"{free_days} Days", "🟡 Configured", "Free days included in carrier quote"],
        ["EPR Recycling Mode", recycling_mode, "🟡 Provision", f"${recycling_rate_per_kwh}/kWh EoL provision rate"],
    ]

    for row_idx, row in enumerate(data2, start=4):
        ws2.append(row)
        for col_idx in range(1, 5):
            ws2.cell(row=row_idx, column=col_idx).border = thin_border

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 15)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------
# Export Button
# ---------------------------------------------------------
st.subheader(t["export_title"])

if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None

if st.button(t["btn_excel"]):
    st.session_state.excel_bytes = generate_excel_bytes()

if st.session_state.excel_bytes:
    st.download_button(
        label="📥 Click Here to Download Excel Report (.xlsx)",
        data=st.session_state.excel_bytes,
        file_name=f"Green_Logistics_Landed_Cost_{clean_hs}.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        use_container_width=True,
    )
